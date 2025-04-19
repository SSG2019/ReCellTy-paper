import os
import pandas as pd
from pycparser.ply.yacc import LRTable
from langchain_neo4j import Neo4jGraph
from langchain_openai import ChatOpenAI
from Agent.name_joint import cell_type_qa
from openpyxl import Workbook, load_workbook
from langchain.chat_models import init_chat_model

os.environ["NEO4J_URI"] = "bolt://localhost:7687"
os.environ["NEO4J_USERNAME"] = "neo4j"
os.environ["NEO4J_PASSWORD"] = "123456abc"
graph = Neo4jGraph()

output_excel = '../experimental_analysis.xlsx'
wb = load_workbook(output_excel)
column = ['Tissue', 'Markers',
          'Exp1_BroadCellType', 'Exp1_SelectedFeatures', 'Exp1_FinalCellType',
          'Exp2_BroadCellType', 'Exp2_SelectedFeatures', 'Exp2_FinalCellType',
          'Exp3_BroadCellType', 'Exp3_SelectedFeatures', 'Exp3_FinalCellType',
          'Exp4_BroadCellType', 'Exp4_SelectedFeatures', 'Exp4_FinalCellType',
          'Exp5_BroadCellType', 'Exp5_SelectedFeatures', 'Exp5_FinalCellType']

input_folder = '../data/Azimuth'
csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]

monitor = pd.read_csv('monitor.csv')
method_num = 0

target_sheet_name = ['gpt-4o-mini', 'gpt-4o', 'deepseek-chat', 'claude-3-7']
for h in range(monitor['model'][method_num], 4):

    if target_sheet_name[h] not in wb.sheetnames:
        ws = wb.create_sheet(title=target_sheet_name[h])
    else:
        ws = wb[target_sheet_name[h]]
    for i in range(17):
        ws.cell(row=2, column=i + 1, value=column[i])
    wb.save(output_excel)

    length_a = 0
    for i in csv_files:
        tissue, _ = i.split("_", 1)
        data = pd.read_csv(input_folder + "/" + i)
        for j in range(len(data)):
            ws.cell(row=3 + length_a + j, column=1, value=tissue)
            ws.cell(row=3 + length_a + j, column=2, value=data['Markers'][j])
        length_a += len(data)
    wb.save(output_excel)

    if target_sheet_name[h] == 'gpt-4o-mini':
        your_api_key_openai = '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'
        os.environ["OPENAI_API_KEY"] = your_api_key_openai
        model = init_chat_model('gpt-4o-mini-2024-07-18', model_provider="openai", temperature=0)
    elif target_sheet_name[h] == 'gpt-4o':
        your_api_key_openai = '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'
        os.environ["OPENAI_API_KEY"] = your_api_key_openai
        model = init_chat_model('gpt-4o-2024-11-20', model_provider="openai", temperature=0)
    elif target_sheet_name[h] == 'deepseek-chat':
        your_api_key_deepseek = 'sk-c1007a49cfc641e18de09e292eb989f5'
        os.environ["OPENAI_API_KEY"] = your_api_key_deepseek
        os.environ["OPENAI_API_BASE"] = 'https://api.deepseek.com'
        model = ChatOpenAI(model=target_sheet_name[h], temperature=0)
    elif target_sheet_name[h] == 'claude-3-7':
        your_api_key_anthropic = 'sk-JNyWaheJNAHox6cr4aD86a6dE043439dB6A98779E51cF0Fb'
        os.environ["ANTHROPIC_API_KEY"] = your_api_key_anthropic
        model = init_chat_model('claude-3-7-sonnet-20250219', model_provider="anthropic", temperature=0,
                                base_url='https://api.cxhao.com')

    for i in range(monitor['repetition'][method_num], 5):
        for j in range(monitor['tissue_num'][method_num], len(csv_files)):
            initial_row = 0
            for ii in range(0, monitor['tissue_num'][method_num]):
                initial_row += len(pd.read_csv(input_folder + "/" + csv_files[ii]))
            initial_row += 3

            data_anno = pd.read_csv(input_folder + "/" + csv_files[j])

            tissue, _ = csv_files[j].split("_", 1)

            if tissue == 'Adipose':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Adipose tissue', 'Adipose tissue')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Bone Marrow':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Bone marrow', 'Bone marrow')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Fetal Development':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Fetal Development', 'Fetal Development', global_search=True)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Heart':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Heart', 'Heart')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Kidney':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Kidney', 'Kidney')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Liver':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Liver', 'Liver')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Lung':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Lung', 'Lung')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Motor Cortex':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Brain', 'Motor Cortex')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Pancreas':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Pancreas', 'Pancreas')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'PBMC':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Blood', 'PBMC')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Tonsil':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    cell_type, broad_type, _, features = cell_type_qa(graph, data_anno['Markers'][x], model, 'Tonsil', 'Tonsil')
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+3, value=broad_type)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+4, value=features)
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['repetition'][method_num]*3+5, value=cell_type)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            monitor.loc[method_num, 'tissue_num'] += 1
            monitor.loc[method_num, 'data_piece_num'] = 0
            monitor.to_csv('monitor.csv', index=False)

        monitor.loc[method_num, 'repetition'] += 1
        monitor.loc[method_num, 'tissue_num'] = 0
        monitor.to_csv('monitor.csv', index=False)
        # break
    monitor.loc[method_num, 'model'] += 1
    monitor.loc[method_num, 'repetition'] = 0
    monitor.to_csv('monitor.csv', index=False)



