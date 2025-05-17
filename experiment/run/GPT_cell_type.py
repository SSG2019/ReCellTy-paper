import os
import pandas as pd
from langchain_openai import ChatOpenAI
from openpyxl import Workbook, load_workbook
from langchain.chat_models import init_chat_model
from langchain_core.prompts.prompt import PromptTemplate
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser

query = """
Identify cell types of {TissueName} cells using the following markers. Only provide the cell type name. Do not show numbers before the name. Some can be a mixture of multiple cell types.
\n {GeneList}
"""
prompt_template = ChatPromptTemplate([
    ("user", query)
])

output_excel = '../experimental_analysis.xlsx'
wb = load_workbook(output_excel)
column = ['Tissue', 'Markers',
          'Exp1_gpt_4o_mini', 'Exp2_gpt_4o_mini', 'Exp3_gpt_4o_mini', 'Exp4_gpt_4o_mini', 'Exp5_gpt_4o_mini', 'Final_gpt_4o_mini',
          'Exp1_gpt_4o', 'Exp2_gpt_4o', 'Exp3_gpt_4o', 'Exp4_gpt_4o', 'Exp5_gpt_4o','Final_gpt_4o',
          'Exp1_deepseek_chat', 'Exp2_deepseek_chat', 'Exp3_deepseek_chat', 'Exp4_deepseek_chat', 'Exp5_deepseek_chat','Final_deepseek_chat',
          'Exp1_claude_3_7', 'Exp2_claude_3_7', 'Exp3_claude_3_7', 'Exp4_claude_3_7', 'Exp5_claude_3_7', 'Final_claude_3_7']

model_name = ['gpt-4o-mini', 'gpt-4o', 'deepseek-chat', 'claude-3-7']

input_folder = '../data/Azimuth'
csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]

monitor = pd.read_csv('monitor.csv')
method_num = 1

target_sheet_name = 'general-purpose-LLMs'

if target_sheet_name not in wb.sheetnames:
    ws = wb.create_sheet(title=target_sheet_name)
else:
    ws = wb[target_sheet_name]
for i in range(26):
    ws.cell(row=2, column=i + 1, value=column[i])
wb.save(output_excel)

length_a = 0
for i in csv_files:
    tissue, _ = i.split("_", 1)
    data = pd.read_csv(input_folder + "/" + i)
    for j in range(len(data)):
        ws.cell(row=3 + length_a + j, column=1, value=tissue)
        ws.cell(row=3 + length_a + j, column=2, value=data['Markers'][j])
    length_a += len(data)
wb.save(output_excel)

for h in range(monitor['model'][method_num], 4):
    if model_name[h] == 'gpt-4o-mini':
        your_api_key_openai = 'your_api_key'
        os.environ["OPENAI_API_KEY"] = your_api_key_openai
        model = init_chat_model('gpt-4o-mini-2024-07-18', model_provider="openai", temperature=0)
    elif model_name[h] == 'gpt-4o':
        your_api_key_openai = 'your_api_key'
        os.environ["OPENAI_API_KEY"] = your_api_key_openai
        model = init_chat_model('gpt-4o-2024-11-20', model_provider="openai", temperature=0)
    elif model_name[h] == 'deepseek-chat':
        your_api_key_deepseek = 'your_api_key'
        os.environ["OPENAI_API_KEY"] = your_api_key_deepseek
        os.environ["OPENAI_API_BASE"] = 'https://api.deepseek.com'
        model = ChatOpenAI(model=model_name[h], temperature=0)
    elif model_name[h] == 'claude-3-7':
        your_api_key_anthropic = 'your_api_key'
        os.environ["ANTHROPIC_API_KEY"] = your_api_key_anthropic
        model = init_chat_model('claude-3-7-sonnet-20250219', model_provider="anthropic", temperature=0)


    chain = prompt_template | model | StrOutputParser()

    for i in range(monitor['repetition'][method_num], 5):
        for j in range(monitor['tissue_num'][method_num], len(csv_files)):
            initial_row = 0
            for ii in range(0, monitor['tissue_num'][method_num]):
                initial_row += len(pd.read_csv(input_folder + "/" + csv_files[ii]))
            initial_row += 3

            data_anno = pd.read_csv(input_folder + "/" + csv_files[j])

            tissue, _ = csv_files[j].split("_", 1)

            if tissue == 'Adipose':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Adipose', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Bone Marrow':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Bone Marrow', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Fetal Development':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Fetal Development', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Heart':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Heart', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Kidney':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Kidney', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Liver':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Liver', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Lung':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Lung', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Motor Cortex':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Motor Cortex', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Pancreas':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Pancreas', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'PBMC':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'PBMC', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            elif tissue == 'Tonsil':
                for x in range(monitor['data_piece_num'][method_num], len(data_anno)):
                    response = chain.invoke({'TissueName': 'Tonsil', 'GeneList': data_anno['Markers'][x]})
                    ws.cell(row=initial_row+monitor['data_piece_num'][method_num], column=monitor['model'][method_num]*6+monitor['repetition'][method_num]+3, value=response)
                    wb.save(output_excel)
                    monitor.loc[method_num, 'data_piece_num'] += 1
                    monitor.to_csv('monitor.csv', index=False)

            monitor.loc[method_num, 'tissue_num'] += 1
            monitor.loc[method_num, 'data_piece_num'] = 0
            monitor.to_csv('monitor.csv', index=False)

        monitor.loc[method_num, 'repetition'] += 1
        monitor.loc[method_num, 'tissue_num'] = 0
        monitor.to_csv('monitor.csv', index=False)

    monitor.loc[method_num, 'model'] += 1
    monitor.loc[method_num, 'repetition'] = 0
    monitor.to_csv('monitor.csv', index=False)