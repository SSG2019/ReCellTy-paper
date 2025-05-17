import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook

def read_excel_sheet(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df

def find_top_strings(df, row_index, target_columns, top_n, feature=False):
    # elif not feature:
    normalised_to_originals = defaultdict(Counter)
    total_counter = Counter()
    # print(df)

    for col in target_columns:
        cell = df.iat[row_index, col]
        print(cell)
        if pd.notnull(cell):
            items = str(cell).split(',')
            for item in items:
                original = item.strip()
                normalised = original.lower()
                total_counter[normalised] += 1
                normalised_to_originals[normalised][original] += 1


    most_common = total_counter.most_common(top_n)

    if not most_common:
        return 'NA'

    norm_word, count = most_common[0]
    original_form = normalised_to_originals[norm_word].most_common(1)[0][0]

    if top_n >= 2:
        top_originals = []
        for norm_word, count in most_common:
            original_form = normalised_to_originals[norm_word].most_common(1)[0][0]
            top_originals.append(original_form)

        original_form = ', '.join(top_originals)

    return original_form


output_excel = '../experimental_analysis.xlsx'
target_sheet_name = ['gpt-4o-mini', 'gpt-4o', 'deepseek-chat', 'claude-3-7']

column_BroadCellType = ['Exp1_BroadCellType', 'Exp2_BroadCellType', 'Exp3_BroadCellType', 'Exp4_BroadCellType', 'Exp5_BroadCellType']
column_SelectedFeatures = ['Exp1_SelectedFeatures', 'Exp2_SelectedFeatures', 'Exp3_SelectedFeatures', 'Exp4_SelectedFeatures', 'Exp5_SelectedFeatures']
column_FinalCellType = ['Exp1_FinalCellType', 'Exp2_FinalCellType', 'Exp3_FinalCellType', 'Exp4_FinalCellType', 'Exp5_FinalCellType']

target_column_name = [column_BroadCellType, column_SelectedFeatures, column_FinalCellType]

for i in target_sheet_name:

    wb = load_workbook(output_excel)
    ws = wb[i]
    df = read_excel_sheet(output_excel, i)

    ws.cell(row=2, column=18, value='BroadCellType')
    ws.cell(row=2, column=19, value='SelectedFeatures')
    ws.cell(row=2, column=20, value='FinalCellType')
    wb.save(output_excel)

    target_columns_BroadCellType = [2, 5, 8, 11, 14]
    target_columns_SelectedFeatures = [3, 6, 9, 12, 15]
    target_columns_FinalCellType = [4, 7, 10, 13, 16]

    if i != 'claude-3-7':
        for j in range(1, 329):

            BroadCellType = find_top_strings(df, j, target_columns_BroadCellType, 1)
            SelectedFeatures = find_top_strings(df, j, target_columns_SelectedFeatures, 3)
            FinalCellType = find_top_strings(df, j, target_columns_FinalCellType, 1)
            ws.cell(row=j+2, column=18, value=BroadCellType)
            ws.cell(row=j+2, column=19, value=SelectedFeatures)
            ws.cell(row=j+2, column=20, value=FinalCellType)

        # print(y)
        # print('**************')
    elif i == 'claude-3-7':
        for j in range(1, 329):
            SelectedFeatures = find_top_strings(df, j, target_columns_SelectedFeatures, 3)
            ws.cell(row=j + 2, column=19, value=SelectedFeatures)


        # print(y)
        # print('**************')

    wb.save(output_excel)







