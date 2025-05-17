import os
import pandas as pd
from pycparser.ply.yacc import LRTable
from langchain_neo4j import Neo4jGraph
from langchain_openai import ChatOpenAI
from Agent.name_joint import cell_type_qa
from openpyxl import Workbook, load_workbook
from langchain.chat_models import init_chat_model

output_excel = '../experimental_analysis.xlsx'
wb = load_workbook(output_excel)
sheet = 'agreement'


input_folder = '../data/Azimuth'
csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]

if sheet not in wb.sheetnames:
    ws = wb.create_sheet(title=sheet)
else:
    ws = wb[sheet]


length_a = 0
for i in csv_files:
    tissue, _ = i.split("_", 1)
    data = pd.read_csv(input_folder + "/" + i)
    for j in range(len(data)):
        ws.cell(row=3 + length_a + j, column=1, value=tissue)
        ws.cell(row=3 + length_a + j, column=2, value=data['Markers'][j])
        ws.cell(row=3 + length_a + j, column=3, value=data['Manual Annotation'][j])
    length_a += len(data)
wb.save(output_excel)